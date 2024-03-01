from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import openpyxl
from datetime import datetime


def getDesiredKeyAndColumn(key_value_dict, desired_key, desired_column):
    # Check if the desired key exists in the dictionary
    if desired_key in key_value_dict:
        # Check if the desired column exists for the desired key
        if desired_column in key_value_dict[desired_key]:
            value_For_Desired_Column = key_value_dict[desired_key][desired_column]
            print(
                f"Value for key '{desired_key}' in column '{desired_column}': {value_For_Desired_Column}"
            )

            return value_For_Desired_Column

        else:
            print(f"column '{desired_column}' not found for key '{desired_key}'")

    else:
        print(f"key '{desired_key}' not found in the dictionary.")


def findElementByXpath(waitDriver, locator):
    try:
        element = waitDriver.until(EC.element_to_be_clickable((By.XPATH, locator)))
    except Exception as E:
        element = waitDriver.until(EC.element_to_be_clickable((By.XPATH, locator)))
    return element


def enterValueInGivenField(
    locator, key_value_dict, uniqueKey, fieldIdentify, waitDriver
):

    element = findElementByXpath(waitDriver, locator)

    element.send_keys(
        str(getDesiredKeyAndColumn(key_value_dict, uniqueKey, fieldIdentify))
    )
    time.sleep(1)


def workBookconfig():
    log_info = "Success"

    # Open the Excel file
    excel_file_path = "files\\Log.xlsx"
    # workbook = openpyxl.load_workbook(excel_file_path)
    sheet_name = "Sheet1"

    try:

        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.get_sheet_by_name(sheet_name)
        print("Excel Found")

    except FileNotFoundError:
        print("Excel not Found")

        # If file doesnt exist,create a new file
        workbook = openpyxl.Workbook()
        workbook.remove(
            workbook.active
        )  # Remove the default sheet created with the new workbook

        # Select the appropriate sheet
        
        sheet = (
            workbook[sheet_name]
            if sheet_name in workbook.sheetnames
            else workbook.create_sheet(sheet_name)
        )

    return workbook, sheet_name, sheet, excel_file_path
